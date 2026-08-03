import os
import io
import re
import zipfile
import warnings
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import seaborn as sns

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
requests.packages.urllib3.disable_warnings()

# 1. Calcular automáticamente el día de ayer en la zona horaria de Perú (GMT-5)
hora_peru = datetime.utcnow() - timedelta(hours=5)
ayer = hora_peru - timedelta(days=1)

dia_str = ayer.strftime("%d")      
mes_num = ayer.strftime("%m")      
anio_str = ayer.strftime("%Y")     

meses_es = {
    "01": "01_Enero", "02": "02_Febrero", "03": "03_Marzo", "04": "04_Abril",
    "05": "05_Mayo", "06": "06_Junio", "07": "07_Julio", "08": "08_Agosto",
    "09": "09_Septiembre", "10": "10_Octubre", "11": "11_Noviembre", "12": "12_Diciembre"
}
mes_carpeta = meses_es[mes_num]

url = (
    f"https://www.coes.org.pe/portal/browser/download?url=Post%20Operaci%C3%B3n"
    f"%2FReportes%2FIEOD%2F{anio_str}%2F{mes_carpeta}%2F{dia_str}%2FAnexoA_{dia_str}{mes_num}.xlsx"
)

filename = f"AnexoA_{dia_str}{mes_num}.xlsx"
print(f"Fecha de procesamiento (Ayer): {ayer.strftime('%Y-%m-%d')}")
print(f"Descargando desde COES: {filename}")

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}

csv_historico = "historico_generacion_rer.csv"
df_consolidado = None

# Manejo seguro de la descarga y extracción
try:
    response = requests.get(url, headers=headers, verify=False, timeout=30)
    if response.status_code != 200:
        print(f"[Aviso] Archivo no disponible en el servidor del COES (Status: {response.status_code}).")
    else:
        content_bytes = io.BytesIO(response.content)
        if zipfile.is_zipfile(content_bytes):
            wb = load_workbook(filename=content_bytes, data_only=True)
            if "GENERACION RER" in wb.sheetnames:
                ws = wb["GENERACION RER"]
                fecha_formateada = ayer.strftime("%Y-%m-%d")
                centrales_detectadas = {}

                for col in range(1, ws.max_column + 1):
                    celda_val = ws.cell(row=7, column=col).value
                    if celda_val:
                        texto_celda = str(celda_val).strip().upper()
                        es_eolica = bool(re.search(r'\bC\.E\.|\bCE\b', texto_celda))
                        es_solar = bool(re.search(r'\bC\.S\.|\bCS\b', texto_celda))
                        
                        if es_eolica or es_solar:
                            tipo_completo = "Eólica" if es_eolica else "Solar"
                            centrales_detectadas[col] = {
                                'nombre': str(celda_val).strip(),
                                'tipo': tipo_completo
                            }

                datos_dia = []
                for row in range(8, 56):
                    intervalo = ws.cell(row=row, column=2).value or ws.cell(row=row, column=1).value
                    registro = {
                        'Fecha': fecha_formateada,
                        'Intervalo': str(intervalo).strip() if intervalo else f"H_{row-7}"
                    }
                    
                    for col, info in centrales_detectadas.items():
                        val = ws.cell(row=row, column=col).value
                        try:
                            registro[info['nombre']] = float(val) if val is not None else 0.0
                        except ValueError:
                            registro[info['nombre']] = 0.0
                        
                    datos_dia.append(registro)

                df_nuevo = pd.DataFrame(datos_dia)

                if not df_nuevo.empty:
                    columnas_activas = ['Fecha', 'Intervalo']
                    for col in df_nuevo.columns:
                        if col in ['Fecha', 'Intervalo']:
                            continue
                        if df_nuevo[col].abs().sum() > 0.01:
                            columnas_activas.append(col)
                    df_nuevo = df_nuevo[columnas_activas]

                if os.path.exists(csv_historico):
                    df_antiguo = pd.read_csv(csv_historico)
                    df_antiguo = df_antiguo[df_antiguo['Fecha'] != fecha_formateada]
                    df_consolidado = pd.concat([df_antiguo, df_nuevo], ignore_index=True)
                else:
                    df_consolidado = df_nuevo

                columnas_ordenadas = ['Fecha', 'Intervalo'] + [c for c in df_consolidado.columns if c not in ['Fecha', 'Intervalo']]
                df_consolidado = df_consolidado[columnas_ordenadas]
                df_consolidado.to_csv(csv_historico, index=False, encoding='utf-8')
                print(f"¡Éxito! Datos guardados en {csv_historico}.")

except Exception as e:
    print(f"[Error en procesamiento]: {e}")

# Si no se pudo descargar hoy pero existe un histórico previo, lo cargamos para poder graficar
if df_consolidado is None and os.path.exists(csv_historico):
    print("Cargando datos históricos existentes para generar la gráfica...")
    df_consolidado = pd.read_csv(csv_historico)

# ==========================================
# 2. GENERACIÓN DEL GRÁFICO PARA EL README
# ==========================================
def generar_grafico_perfil(df):
    if df is None or df.empty:
        print("[Aviso] No hay datos disponibles para graficar.")
        fig, ax = plt.subplots(figsize=(6, 2))
        ax.text(0.5, 0.5, "Sin datos disponibles", ha='center', va='center')
        ax.axis('off')
        plt.savefig("perfil_generacion_rer.png", dpi=100)
        plt.close()
        return

    sns.set_theme(style="whitegrid")
    centrales = [c for c in df.columns if c not in ['Fecha', 'Intervalo']]
    num_centrales = len(centrales)
    
    if num_centrales == 0:
        return

    cols = 3
    rows = (num_centrales + cols - 1) // cols
    fig, axes = plt.subplots(rows, cols, figsize=(18, 4 * rows), sharex=True, sharey=False)
    
    if num_centrales == 1:
        axes = [axes]
    else:
        axes = axes.flatten()

    fechas = df['Fecha'].unique()
    ultima_fecha = max(fechas)

    for i, central in enumerate(centrales):
        ax = axes[i]
        
        # Perfil Histórico (gris)
        for fecha in fechas:
            if fecha == ultima_fecha:
                continue
            df_dia = df[df['Fecha'] == fecha]
            ax.plot(df_dia['Intervalo'], df_dia[central], color='gray', alpha=0.2, linewidth=1)

        # Último día (resaltado en naranja)
        df_ultimo = df[df['Fecha'] == ultima_fecha]
        ax.plot(
            df_ultimo['Intervalo'], 
            df_ultimo[central], 
            color='#d95f02', 
            linewidth=2.5, 
            label=f'Último día ({ultima_fecha})'
        )

        ax.set_title(f"{central}", fontsize=10, fontweight='bold')
        ax.set_ylabel("MW", fontsize=8)
        ax.tick_params(axis='x', rotation=90, labelsize=6)
        ax.grid(True, linestyle='--', alpha=0.5)
        ax.legend(loc='upper right', fontsize=8)

    for j in range(i + 1, len(axes)):
        fig.delaxes(axes[j])

    plt.suptitle("Perfil Diario de Generación RER", fontsize=15, fontweight='bold', y=1.01)
    plt.tight_layout()
    plt.savefig("perfil_generacion_rer.png", dpi=200, bbox_inches='tight')
    plt.close()
    print("Gráfico 'perfil_generacion_rer.png' generado exitosamente.")

# Ejecución de la función de graficación
generar_grafico_perfil(df_consolidado)
